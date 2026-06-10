"""Engine tests: detection, output sheets, and original-data integrity."""
from __future__ import annotations

import datetime as _dt
import os
import shutil

import pytest
from openpyxl import Workbook, load_workbook

from core.models import AnalysisOptions, ColumnType
from core.pipeline import Engine

from .make_sample import make_sample


@pytest.fixture()
def sample(tmp_path):
    p = tmp_path / "sample.xlsx"
    make_sample(str(p))
    return str(p)


def test_profile_detects_types(sample):
    eng = Engine()
    prof = eng.profile(sample)
    assert prof.primary.sheet_name == "Sales Data"
    assert prof.primary.row_count == 400
    types = {c.name: c.ctype for c in prof.primary.columns}
    assert types["Revenue"] == ColumnType.CURRENCY
    assert types["Margin %"] == ColumnType.PERCENT
    assert types["Date"] == ColumnType.DATE
    assert types["Region"] == ColumnType.CATEGORICAL


def test_multi_sheet_detection(sample):
    prof = Engine().profile(sample)
    names = {t.sheet_name for t in prof.tables}
    assert {"Sales Data", "Targets"} <= names


def test_run_creates_sheets_and_keeps_original(sample):
    wb0 = load_workbook(sample)
    orig = {n: load_workbook(sample)[n]["A1"].value for n in wb0.sheetnames}
    wb0.close()

    res = Engine().run(sample, AnalysisOptions())
    assert set(res.sheets_created) == {
        "KPI Analysis", "Pivot Analysis", "Dashboard", "Executive Summary"}

    wb1 = load_workbook(sample)
    for name, a1 in orig.items():
        assert name in wb1.sheetnames, f"original sheet {name} lost"
        assert wb1[name]["A1"].value == a1, f"original sheet {name} modified"
    wb1.close()


def test_offline_summary_fallback(sample):
    res = Engine(narrator=None).run(sample, AnalysisOptions(executive_summary=True,
                                                            dashboard=False, pivot=False,
                                                            kpi=False))
    assert res.summary_used_llm is False
    assert any("metrics" in n for n in res.notes)


def test_rerun_regenerates_cleanly(sample):
    Engine().run(sample, AnalysisOptions())
    res2 = Engine().run(sample, AnalysisOptions())
    # A second run regenerates the analysis sheets in place (no "(2)" duplicates).
    assert "(2)" not in " ".join(res2.sheets_created)
    wb = load_workbook(sample)
    for base in ("KPI Analysis", "Dashboard", "Executive Summary"):
        assert sum(1 for n in wb.sheetnames if n.split(" (")[0] == base) == 1
    wb.close()


def test_no_data_workbook_raises(tmp_path):
    p = tmp_path / "empty.xlsx"
    wb = Workbook(); wb.active["A1"] = "just a title"; wb.save(str(p))
    with pytest.raises(ValueError):
        Engine().profile(str(p))
