"""Ingest the wide "code / description / category-banner" account catalog.

The hospital's chart-of-accounts catalog is laid out in column GROUPS: each pair
of columns is (code, description), and the description column's HEADER is a
category banner — "REVENUES CODES", "PURCHASES", "GENERAL EXPENSES", "SALARIES",
… — that classifies every code beneath it. This teaches the library not just
what each account is called, but whether it is revenue, an expense, cash, etc.,
which lets the engine infer what a workbook is ABOUT.

Run:  python tools/ingest_account_categories.py "header with category and description.xlsx" [--map account]
"""
from __future__ import annotations

import argparse
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook

from core.library.store import CodeMap, get_library, norm_code, save_library


def _clean_category(banner: str) -> str:
    """'REVENUES CODES' -> 'revenues'; 'GENERAL EXPENSES' -> 'general expenses'."""
    s = re.sub(r"\s+", " ", str(banner or "").strip().lower())
    s = re.sub(r"\bcodes?\b", "", s).strip()        # drop a trailing/leading "code(s)"
    return s or "uncategorized"


def ingest(path: str, map_name: str = "account") -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return {"added": 0, "categories": {}}

    header = rows[0]
    # Find (code_col, desc_col) pairs: a code column's header starts with "CODE",
    # and the next column's header is the category banner.
    pairs: list[tuple[int, int, str]] = []
    for c in range(len(header) - 1):
        h = str(header[c] or "").strip().upper()
        if h.startswith("CODE"):
            banner = header[c + 1]
            if banner:
                pairs.append((c, c + 1, _clean_category(banner)))

    lib = get_library(refresh=True)
    cm = lib.code_maps.get(map_name) or CodeMap(name=map_name, label=map_name)
    added = 0
    cat_counts: dict[str, int] = {}
    for code_c, desc_c, category in pairs:
        for row in rows[1:]:
            if code_c >= len(row) or desc_c >= len(row):
                continue
            code = norm_code(row[code_c])
            desc = str(row[desc_c]).strip() if row[desc_c] else ""
            if not code or not desc:
                continue
            cm.entries[code] = desc
            cm.categories[code] = category
            cat_counts[category] = cat_counts.get(category, 0) + 1
            added += 1
    cm.reindex()
    lib.code_maps[map_name] = cm
    lib.meta.setdefault("sources", []).append(
        {"file": path.split("\\")[-1], "kind": "account-categories",
         "category": map_name, "codes": added})
    save_library(lib)
    return {"added": added, "categories": cat_counts}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--map", default="account")
    a = ap.parse_args()
    res = ingest(a.path, a.map)
    print(f"Ingested {res['added']} codes into '{a.map}'.")
    print("Categories:")
    for k, v in sorted(res["categories"].items(), key=lambda kv: -kv[1]):
        print(f"  {v:5d}  {k}")
