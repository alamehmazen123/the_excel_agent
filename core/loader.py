"""Open a workbook and auto-detect its data tables.

Strategy per worksheet:
  1. Find the densest contiguous block of cells (the data region).
  2. Pick the header row: the first mostly-text row at the top of that block
     that is followed by typed data.
  3. Build a TableProfile of {column_name: value} row dicts and profile it.

We deliberately keep this tolerant: messy real-world files have title rows,
blank spacer rows, and multiple tables. We grab the largest well-formed table
per sheet and treat the biggest overall as the workbook's primary table.
"""
from __future__ import annotations

import datetime as _dt
import os
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .constants import output_sheet_names
from .models import TableProfile, WorkbookProfile
from .pivot_detect import detect_pivot_sheets
from .profiler import profile_table


def _cell_is_empty(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _detect_header_grid(grid: list[list[Any]], ncols: int) -> Optional[int]:
    """Find the header row (0-based index into ``grid``) from in-memory values."""
    need = max(2, ncols * 0.5)
    limit = min(len(grid), 25)
    for r in range(limit):
        non_empty = [v for v in grid[r] if not _cell_is_empty(v)]
        if len(non_empty) < need:
            continue
        text_like = sum(1 for v in non_empty if isinstance(v, str))
        # A header row is mostly strings, and the row below it has some data.
        if text_like >= len(non_empty) * 0.6:
            if r + 1 < len(grid) and any(not _cell_is_empty(v) for v in grid[r + 1]):
                return r
    return None


def _dedupe_headers(raw: list[Any], min_col: int) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for i, v in enumerate(raw):
        name = str(v).strip() if not _cell_is_empty(v) else \
            f"Column_{get_column_letter(min_col + i)}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        names.append(name)
    return names


def _extract_table(ws) -> Optional[TableProfile]:
    """Stream the sheet ONCE (O(rows)) and build a profiled table.

    openpyxl read-only mode re-streams from the top on every ``ws.cell()`` call,
    which is O(rows²). Iterating with ``ws.iter_rows()`` a single time avoids that
    and keeps loading fast even for large sheets.
    """
    grid: list[list[Any]] = []
    fmt_by_col: dict[int, str] = {}   # 1-based col index -> source number format
    ncols = 0
    for row_cells in ws.iter_rows():
        vals: list[Any] = []
        for idx, cell in enumerate(row_cells, start=1):
            v = cell.value
            vals.append(v)
            # Capture the format from the first numeric/date cell of each column
            # (skip header text), so currency/percent detection is data-driven.
            if (idx not in fmt_by_col
                    and isinstance(v, (int, float, _dt.date, _dt.datetime))
                    and not isinstance(v, bool)):
                fmt_by_col[idx] = getattr(cell, "number_format", "General") or "General"
        if len(vals) > ncols:
            ncols = len(vals)
        grid.append(vals)

    if len(grid) < 2 or ncols == 0:
        return None
    for vals in grid:                 # pad ragged rows to a uniform width
        if len(vals) < ncols:
            vals.extend([None] * (ncols - len(vals)))

    hidx = _detect_header_grid(grid, ncols)
    if hidx is None:
        return None

    raw_headers = list(grid[hidx][:ncols])
    while raw_headers and _cell_is_empty(raw_headers[-1]):
        raw_headers.pop()
    if not raw_headers:
        return None
    ncols = len(raw_headers)
    names = _dedupe_headers(raw_headers, 1)

    rows: list[dict[str, Any]] = []
    last_data = hidx
    for r in range(hidx + 1, len(grid)):
        values = grid[r][:ncols]
        if all(_cell_is_empty(v) for v in values):
            continue
        rows.append({names[i]: values[i] for i in range(ncols)})
        last_data = r
    if not rows:
        return None

    fmt_by_name = {names[i]: fmt_by_col.get(i + 1, "General") for i in range(ncols)}

    table = TableProfile(
        sheet_name=ws.title, header_row=hidx + 1,
        first_data_row=hidx + 2, last_data_row=last_data + 1,
        first_col=1, last_col=ncols, rows=rows,
    )
    profile_table(table, fmt_by_name)   # formats drive currency/percent detection
    return table


def load_workbook_profile(path: str) -> WorkbookProfile:
    """Open ``path`` and return a fully profiled :class:`WorkbookProfile`."""
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Workbook not found: {path}")

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # password-protected, corrupt, wrong format
        raise ValueError(
            "Could not open this workbook. It may be password-protected, "
            "corrupted, or not a valid .xlsx file."
        ) from exc

    profile = WorkbookProfile(path=path, sheet_names=list(wb.sheetnames))
    output_names = set(output_sheet_names())

    # Sheets that already hold a PivotTable are left completely untouched.
    pivot_sheets = set(detect_pivot_sheets(path))
    profile.pivot_sheets = [s for s in wb.sheetnames if s in pivot_sheets]

    for ws in wb.worksheets:
        if ws.title in output_names:
            # Skip sheets we previously produced so re-runs don't analyze them.
            continue
        if ws.title in pivot_sheets:
            # Already contains a pivot table -> do not analyze or modify it.
            continue
        try:
            table = _extract_table(ws)
        except Exception as exc:  # noqa: BLE001 - one bad sheet shouldn't kill the run
            profile.warnings.append(f"Skipped sheet '{ws.title}': {exc}")
            continue
        if table is not None and table.measures:
            profile.tables.append(table)
        elif table is not None:
            # Keep tables without measures too (still useful for counts/pivots).
            profile.tables.append(table)

    wb.close()

    if not profile.tables:
        raise ValueError(
            "No analyzable data tables were found in this workbook. "
            "Make sure at least one sheet has a header row and rows of data."
        )

    # Primary table = the one with the most data cells.
    profile.primary_table_index = max(
        range(len(profile.tables)),
        key=lambda i: profile.tables[i].row_count * len(profile.tables[i].columns),
    )
    return profile
