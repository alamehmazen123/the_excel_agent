"""Render SheetSpecs into the workbook using openpyxl.

Contract: the original sheets are never modified -- we only append new sheets.
Output is written to a temp file and atomically swapped in to avoid leaving a
corrupted workbook if the process is interrupted.
"""
from __future__ import annotations

import os
import tempfile
from typing import Optional

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .constants import output_sheet_names

from .render import (ChartKind, ChartSpec, DataTable, NumberFormat, SheetSpec,
                     TextBlock)

# Palette
_NAVY = "1F3864"
_BLUE = "2E5496"
_LIGHT = "D6E0F0"
_GREEN = "2E7D32"
_RED = "C62828"
_GREY = "595959"
_WHITE = "FFFFFF"

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_NUMFMT = {
    NumberFormat.GENERAL: "General",
    NumberFormat.INTEGER: "#,##0",
    NumberFormat.DECIMAL: "#,##0.00",
    NumberFormat.CURRENCY: "$#,##0.00",
    NumberFormat.LBP: '#,##0" LBP"',
    NumberFormat.PERCENT: "0.0%",
    NumberFormat.DATE: "yyyy-mm-dd",
}


def _remove_existing_outputs(wb) -> None:
    """Delete any analysis sheets from a previous run so we regenerate cleanly."""
    targets = set(output_sheet_names())
    for ws in list(wb.worksheets):
        name = ws.title
        # match exact names and versioned variants like "Dashboard (2)"
        base = name.split(" (")[0]
        if base in targets and len(wb.worksheets) > 1:
            del wb[name]


def _unique_sheet_name(wb, base: str) -> str:
    if base not in wb.sheetnames:
        return base
    i = 2
    while f"{base} ({i})" in wb.sheetnames:
        i += 1
    return f"{base} ({i})"


_BRAND_BLUE = "0070C0"   # vivid blue for the heading / hospital name


def _write_heading(ws: Worksheet, spec: SheetSpec, row: int) -> int:
    # Heading text rendered BOLD + BLUE on white (e.g. "SAHEL GENERAL HOSPITAL").
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=spec.heading)
    c.font = Font(name="Calibri", size=20, bold=True, color=_BRAND_BLUE)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 32
    row += 1
    if spec.subheading:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        s = ws.cell(row=row, column=1, value=spec.subheading)
        s.font = Font(size=11, bold=True, color=_NAVY)
        s.alignment = Alignment(indent=1)
        row += 1
    return row + 1


def _write_kpi_tiles(ws: Worksheet, spec: SheetSpec, row: int) -> int:
    col = 1
    span = 2          # each tile spans 2 columns, 3 rows
    for tile in spec.kpi_tiles:
        # Label
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        lc = ws.cell(row=row, column=col, value=tile.label.upper())
        lc.font = Font(size=9, bold=True, color=_WHITE)
        lc.fill = PatternFill("solid", fgColor=_BLUE)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        # Value
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + span - 1)
        vc = ws.cell(row=row + 1, column=col, value=tile.value)
        vc.font = Font(size=16, bold=True, color=_NAVY)
        vc.fill = PatternFill("solid", fgColor=_LIGHT)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        # Caption
        ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + span - 1)
        cap_color = _GREY if tile.good is None else (_GREEN if tile.good else _RED)
        cc = ws.cell(row=row + 2, column=col, value=tile.caption)
        cc.font = Font(size=9, color=cap_color)
        cc.fill = PatternFill("solid", fgColor=_LIGHT)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row + 1].height = 24
        col += span
        if col > 8:       # wrap to next tile row after 4 tiles
            col = 1
            row += 4
    # advance past the last (possibly partial) tile row
    return row + (4 if col != 1 else 0) + 1


def _write_table(ws: Worksheet, table: DataTable, row: int) -> tuple[int, dict]:
    """Render a table; return (next_row, anchor info for charting)."""
    tc = ws.cell(row=row, column=1, value=table.title)
    tc.font = Font(size=12, bold=True, color=_NAVY)
    row += 1
    header_row = row
    n_cols = len(table.headers)
    for j, h in enumerate(table.headers, start=1):
        hc = ws.cell(row=row, column=j, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_BLUE)
        hc.alignment = Alignment(horizontal="center")
        hc.border = _BORDER
    row += 1
    first_data = row
    for r in table.rows:
        for j, val in enumerate(r, start=1):
            cell = ws.cell(row=row, column=j, value=val)
            cell.border = _BORDER
            if table.formats and j - 1 < len(table.formats):
                cell.number_format = _NUMFMT[table.formats[j - 1]]
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F5FB")
        row += 1
    info = {
        "header_row": header_row, "first_data": first_data,
        "last_data": row - 1, "n_cols": n_cols,
    }
    _apply_table_visuals(ws, table, info)
    return row + 1, info


# Data-bar / color-scale palette (Smart Tables 2.0).
_BAR_COLOR = "5B9BD5"        # soft blue gradient bar
_SCALE_LOW = "F8696B"        # red (low)
_SCALE_MID = "FFEB84"        # amber (mid)
_SCALE_HIGH = "63BE7B"       # green (high)


def _apply_table_visuals(ws: Worksheet, table: DataTable, info: dict) -> None:
    """Add in-cell data bars and/or a green→red color scale to chosen columns,
    so totals read at a glance without opening a chart."""
    if info["last_data"] < info["first_data"]:
        return
    for idx in getattr(table, "bar_columns", []):
        if not 0 <= idx < info["n_cols"]:
            continue
        col = get_column_letter(idx + 1)
        rng = f"{col}{info['first_data']}:{col}{info['last_data']}"
        try:
            ws.conditional_formatting.add(rng, DataBarRule(
                start_type="min", end_type="max", color=_BAR_COLOR))
        except Exception:
            pass
    for idx in getattr(table, "scale_columns", []):
        if not 0 <= idx < info["n_cols"]:
            continue
        col = get_column_letter(idx + 1)
        rng = f"{col}{info['first_data']}:{col}{info['last_data']}"
        try:
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color=_SCALE_LOW,
                mid_type="percentile", mid_value=50, mid_color=_SCALE_MID,
                end_type="max", end_color=_SCALE_HIGH))
        except Exception:
            pass


def _add_chart(ws: Worksheet, chart: ChartSpec, anchor: str,
               data_start_row: int) -> None:
    """Write the chart's data to a helper area and attach a native chart."""
    # Place chart source data in columns J/K (10/11), out of the main view.
    cat_col, val_col = 10, 11
    ws.cell(row=data_start_row, column=cat_col, value="Category")
    ws.cell(row=data_start_row, column=val_col, value=chart.series_name)
    for i, (cat, val) in enumerate(zip(chart.categories, chart.values), start=1):
        ws.cell(row=data_start_row + i, column=cat_col, value=cat)
        ws.cell(row=data_start_row + i, column=val_col, value=val)
    n = len(chart.values)
    if n == 0:
        return

    # Pareto / combo carry an overlay line in columns L (12); write it too.
    line_col = 12
    has_line = chart.kind in (ChartKind.PARETO, ChartKind.COMBO) and chart.line_values
    if has_line:
        ws.cell(row=data_start_row, column=line_col,
                value=chart.line_name or "Cumulative %")
        for i, lv in enumerate(chart.line_values, start=1):
            ws.cell(row=data_start_row + i, column=line_col, value=lv)

    if chart.kind in (ChartKind.BAR, ChartKind.PARETO, ChartKind.COMBO):
        ch = BarChart(); ch.type = "col"
    elif chart.kind == ChartKind.LINE:
        ch = LineChart()
    else:
        ch = PieChart()
    ch.title = chart.title          # title sits ABOVE the chart by default
    ch.height = 7.5
    ch.width = 15

    data_ref = Reference(ws, min_col=val_col, min_row=data_start_row,
                         max_row=data_start_row + n)
    cats_ref = Reference(ws, min_col=cat_col, min_row=data_start_row + 1,
                         max_row=data_start_row + n)
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)

    # --- req 8: chart formatting ---
    if chart.kind in (ChartKind.BAR, ChartKind.LINE):
        ch.legend = None                                   # LEGEND = None
        # show both primary axes
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        ch.x_axis.majorTickMark = "out"
        ch.y_axis.majorTickMark = "out"
        # keep category (date) labels readable
        ch.x_axis.tickLblPos = "low"
    if chart.kind == ChartKind.BAR:
        # Data labels = Outside End, value shown.
        ch.dataLabels = DataLabelList()
        ch.dataLabels.showVal = True
        ch.dataLabels.showSerName = False
        ch.dataLabels.showCatName = False
        ch.dataLabels.showLegendKey = False
        ch.dataLabels.dLblPos = "outEnd"

    # Overlay a line on a secondary axis for Pareto (cumulative %) / combo (trend).
    if has_line:
        line = LineChart()
        line_ref = Reference(ws, min_col=line_col, min_row=data_start_row,
                             max_row=data_start_row + len(chart.line_values))
        line.add_data(line_ref, titles_from_data=True)
        line.set_categories(cats_ref)
        # Put the line on its own axis so its scale (e.g. 0-100%) is independent.
        line.y_axis.axId = 200
        line.y_axis.crosses = "max"
        ch.y_axis.crosses = "autoZero"
        ch += line
        ch.legend = None
    ws.add_chart(ch, anchor)


def _write_text(ws: Worksheet, block: TextBlock, row: int) -> int:
    style = getattr(block, "style", "normal")
    _AMBER = "C05621"
    # Title styling per block style.
    tc = ws.cell(row=row, column=1, value=block.title)
    if style == "recommend":
        tc.font = Font(size=13, bold=True, color=_RED, underline="single")
    elif style == "warn":
        tc.font = Font(size=12, bold=True, color=_AMBER, underline="single")
    elif style == "highlight":
        tc.font = Font(size=12, bold=True, color=_NAVY)
    else:
        tc.font = Font(size=11, bold=True, color=_NAVY)
    row += 1

    for para in block.paragraphs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        text = para
        if style == "recommend":
            text = f"➤  {para}"
            font = Font(size=11, bold=True, color=_RED)
        elif style == "warn":
            text = f"⚠  {para}"
            font = Font(size=11, bold=True, color=_AMBER)
        elif style == "highlight":
            text = f"•  {para}"
            font = Font(size=11, bold=True, color=_NAVY)
        else:
            font = Font(size=10, color=_GREY)
        pc = ws.cell(row=row, column=1, value=text)
        pc.font = font
        pc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(15, 15 * (len(text) // 90 + 1))
        row += 1
    return row + 1


def _add_listobject(wb, ws: Worksheet, info: dict) -> None:
    """Register the cell block as a real Excel Table (ListObject).

    This lets the Excel COM pass find every table generically (e.g. to apply
    conditional formatting). A Table's DataBodyRange excludes the header row and
    any totals row, which is exactly what 'top-1 per column excluding total' needs.
    """
    if info["last_data"] < info["first_data"] or info["n_cols"] < 1:
        return
    ref = (f"{get_column_letter(1)}{info['header_row']}:"
           f"{get_column_letter(info['n_cols'])}{info['last_data']}")
    # Unique, valid display name across the whole workbook.
    existing = {t for s in wb.worksheets for t in getattr(s, "tables", {})}
    base = f"Tbl_{''.join(ch for ch in ws.title if ch.isalnum()) or 'S'}"
    name, k = base, 1
    while name in existing:
        k += 1
        name = f"{base}{k}"
    try:
        tbl = Table(displayName=name, ref=ref)
        # No visible table style -> keep our custom header/stripe formatting.
        tbl.tableStyleInfo = TableStyleInfo(
            name="", showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False)
        ws.add_table(tbl)
    except Exception:
        pass   # never let table registration break sheet creation


def render_sheet(wb, spec: SheetSpec) -> str:
    """Create and populate one worksheet from ``spec``. Returns its final name."""
    name = _unique_sheet_name(wb, spec.name)
    ws = wb.create_sheet(title=name)
    ws.sheet_view.showGridLines = False
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16

    row = 1
    row = _write_heading(ws, spec, row)
    if spec.kpi_tiles:
        row = _write_kpi_tiles(ws, spec, row)
    for block in spec.text_blocks:
        row = _write_text(ws, block, row)

    chart_helper_row = 1
    for i, table in enumerate(spec.tables):
        row, info = _write_table(ws, table, row)
        _add_listobject(wb, ws, info)
        # Attach a matching chart (if provided by the analyzer) beside the data.
        if i < len(spec.charts):
            anchor = f"M{max(2, row - 12)}"
            _add_chart(ws, spec.charts[i], anchor, chart_helper_row)
            chart_helper_row += len(spec.charts[i].values) + 3

    # Charts with no paired table (e.g. dashboard) get stacked at the bottom.
    for j in range(len(spec.tables), len(spec.charts)):
        anchor = f"A{row}"
        _add_chart(ws, spec.charts[j], anchor, chart_helper_row)
        chart_helper_row += len(spec.charts[j].values) + 3
        row += 16
    return name


_HELPER_SUFFIX = " (Name)"


def inject_hidden_helpers(wb, table, decodes, library) -> list[str]:
    """Write decoded-name helper columns (HIDDEN) next to their code columns on
    the data sheet. Idempotent: an existing same-named helper column is reused
    and overwritten rather than duplicated. Returns the helper names written.

    Decoding is done by reading each code cell straight from the sheet, so it is
    robust to the loader having skipped blank rows (no index-alignment needed).
    """
    if not decodes:
        return []
    ws = wb[table.sheet_name] if table.sheet_name in wb.sheetnames else None
    if ws is None:
        return []

    header_row = table.header_row
    # Map existing header text -> column index so re-runs overwrite in place.
    existing: dict[str, int] = {}
    max_used = 0
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None and str(val) != "":
            max_used = col
            existing[str(val)] = col
    next_col = max_used + 1

    written: list[str] = []
    for dc in decodes:
        src_idx = existing.get(dc.source_name)
        if src_idx is None:
            # Source column not located on the sheet -> skip safely.
            continue
        target_col = existing.get(dc.helper_name)
        if target_col is None:
            target_col = next_col
            next_col += 1
        # Header + decoded values for every data row on the sheet.
        ws.cell(row=header_row, column=target_col, value=dc.helper_name)
        for r in range(table.first_data_row, table.last_data_row + 1):
            raw = ws.cell(row=r, column=src_idx).value
            name = library.decode(dc.cmap_name, raw) if raw not in (None, "") else None
            ws.cell(row=r, column=target_col, value=name)
        ws.column_dimensions[get_column_letter(target_col)].hidden = True
        existing[dc.helper_name] = target_col
        written.append(dc.helper_name)
    return written


def _atomic_save(wb, source_path: str) -> None:
    folder = os.path.dirname(os.path.abspath(source_path))
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=folder)
    os.close(fd)
    try:
        wb.save(tmp)
        wb.close()
        os.replace(tmp, source_path)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise


def inject_positive_helpers(wb, table, helpers: list) -> list[str]:
    """Write HIDDEN sign-flipped POSITIVE helper columns next to revenue money
    columns (e.g. ``ORG_AMOUNT`` → hidden ``ORG_AMOUNT (+)``). The original
    column is untouched; the COM PivotTables aggregate the positive helper so
    they read naturally. ``helpers`` is ``[(source_col, helper_col), …]``."""
    if not helpers:
        return []
    ws = wb[table.sheet_name] if table.sheet_name in wb.sheetnames else None
    if ws is None:
        return []
    header_row = table.header_row
    existing: dict[str, int] = {}
    max_used = 0
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None and str(val) != "":
            max_used = col
            existing[str(val)] = col
    next_col = max_used + 1

    written: list[str] = []
    for source_name, helper_name in helpers:
        src_idx = existing.get(source_name)
        if src_idx is None:
            continue
        target_col = existing.get(helper_name, next_col)
        if helper_name not in existing:
            next_col += 1
        src_fmt = ws.cell(row=table.first_data_row, column=src_idx).number_format
        ws.cell(row=header_row, column=target_col, value=helper_name)
        for r in range(table.first_data_row, table.last_data_row + 1):
            raw = ws.cell(row=r, column=src_idx).value
            cell = ws.cell(row=r, column=target_col)
            cell.value = -raw if isinstance(raw, (int, float)) and not isinstance(raw, bool) else raw
            cell.number_format = src_fmt
        ws.column_dimensions[get_column_letter(target_col)].hidden = True
        existing[helper_name] = target_col
        written.append(helper_name)
    return written


def write_results(source_path: str, specs: list[SheetSpec],
                  inject: Optional[tuple] = None,
                  value_helpers: Optional[tuple] = None) -> list[str]:
    """Regenerate the analysis sheets (removes prior output sheets first).

    ``inject`` (optional) is ``(table, decodes, library)``: hidden decoded-name
    helper columns are written onto the data sheet so every analysis (incl. the
    active PivotTables built afterwards) can group by readable names. Original
    columns are never modified — only new hidden columns are appended.

    Returns the list of created sheet names.
    """
    try:
        wb = load_workbook(source_path)   # full fidelity (keeps formulas/values)
    except Exception as exc:
        raise ValueError(f"Could not open workbook for writing: {exc}") from exc

    if inject is not None:
        table, decodes, library = inject
        try:
            inject_hidden_helpers(wb, table, decodes, library)
        except Exception:
            pass   # decoration is best-effort; never block sheet creation

    if value_helpers is not None:
        vtable, vhelpers = value_helpers
        try:
            inject_positive_helpers(wb, vtable, vhelpers)
        except Exception:
            pass

    _remove_existing_outputs(wb)
    created = [render_sheet(wb, s) for s in specs if s is not None]
    _order_output_sheets(wb)
    _atomic_save(wb, source_path)
    return created


def _order_output_sheets(wb) -> None:
    """Arrange the analysis tabs in the canonical order AFTER the data sheet(s):
    KPI → Pivot → Smart Tables → Insights → Executive Summary → Dashboard.
    Tab order only — sheet contents (including the original data) are untouched.
    Sheets that weren't produced are simply skipped."""
    from .constants import ordered_output_layout  # noqa: PLC0415
    for base in ordered_output_layout():
        for ws in wb.worksheets:
            if ws.title.split(" (")[0] == base:
                try:
                    # Move this output sheet to the very end; doing this for each
                    # sheet in order leaves them sequenced at the tail, with the
                    # data sheet(s) untouched at the front.
                    wb.move_sheet(ws, len(wb.worksheets) - 1 - wb.worksheets.index(ws))
                except Exception:
                    pass
                break


def append_sheets(source_path: str, specs: list[SheetSpec]) -> list[str]:
    """Add sheets WITHOUT removing the other analysis sheets. Only a same-named
    sheet (e.g. an empty 'Pivot Analysis' left by a failed Excel step) is
    replaced. Used for the static pivot fallback so KPI/Dashboard/Summary stay."""
    try:
        wb = load_workbook(source_path)
    except Exception as exc:
        raise ValueError(f"Could not open workbook for writing: {exc}") from exc

    created: list[str] = []
    for spec in specs:
        if spec is None:
            continue
        for ws in list(wb.worksheets):
            if ws.title.split(" (")[0] == spec.name and len(wb.worksheets) > 1:
                del wb[ws.title]
        created.append(render_sheet(wb, spec))
    _order_output_sheets(wb)            # keep the canonical tab order after fallback
    _atomic_save(wb, source_path)
    return created
