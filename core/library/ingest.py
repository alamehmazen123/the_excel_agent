"""Ingest the user's Excel reference files into the library.

The user provides, on several occasions, two kinds of workbook:

* **Glossary files** -- a list of abbreviated headers and their real meaning
  (and optionally a category linking the header to a code domain).
* **Code files** -- a list of codes and their definitions for ONE domain
  (guarantor, department, supplier, doctor, ...). The domain (category) comes
  from the sheet name or is passed explicitly.

Ingestion is incremental and idempotent: re-ingesting updates existing keys and
adds new ones; it never drops knowledge already in the library.

Column detection is heuristic and intentionally forgiving, because each file
the hospital sends may label its columns differently. Adjust ``_HEADER_KEY_HINTS``
/ ``_MEANING_HINTS`` / ``_CODE_HINTS`` / ``_DEF_HINTS`` as real files arrive.
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any, Optional

from openpyxl import load_workbook

from .store import (CodeMap, HeaderEntry, Library, load_library, norm_code,
                    norm_header, save_library)

# Column-name hints (normalized, substring match) used to find the right columns.
_HEADER_KEY_HINTS = ("ABBREV", "ABBREVIATION", "HEADER", "SYMBOL", "CODE", "SHORT", "FIELD")
# Abbreviation-specific hints used ONLY to tell a glossary from a codes file:
# bare "CODE"/"ID" is excluded here because it appears in both kinds.
_GLOSSARY_KEY_HINTS = ("ABBREV", "ABBREVIATION", "HEADER", "SYMBOL", "SHORT", "FIELD")
_MEANING_HINTS = ("MEANING", "DEFINITION", "DESCRIPTION", "FULL", "REAL", "LABEL", "NAME")
_CATEGORY_HINTS = ("CATEGORY", "DOMAIN", "TYPE", "GROUP")
_CODE_HINTS = ("CODE", "ID", "NO", "NUMBER", "KEY")
_DEF_HINTS = ("DEFINITION", "DESCRIPTION", "MEANING", "NAME", "LABEL", "TITLE", "VALUE")


@dataclass
class IngestReport:
    kind: str = ""
    category: str = ""
    source: str = ""
    headers_added: int = 0
    codes_added: int = 0
    sheets: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def summary(self) -> str:
        bits = [f"kind={self.kind or '?'}"]
        if self.category:
            bits.append(f"category={self.category}")
        if self.headers_added:
            bits.append(f"+{self.headers_added} headers")
        if self.codes_added:
            bits.append(f"+{self.codes_added} codes")
        return ", ".join(bits)


def _rows_of(ws) -> list[list[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_header_row(rows: list[list[Any]]) -> int:
    """Index of the first non-empty row (assumed to be the column titles)."""
    for i, row in enumerate(rows):
        if any(c is not None and str(c).strip() for c in row):
            return i
    return 0


def _pick_columns(titles: list[str], *hint_groups: tuple) -> list[Optional[int]]:
    """For each hint group, return the index of the best-matching title column."""
    norm = [norm_header(t) for t in titles]
    picks: list[Optional[int]] = []
    used: set[int] = set()
    for hints in hint_groups:
        chosen: Optional[int] = None
        # exact-ish: title contains a hint word
        for i, t in enumerate(norm):
            if i in used or not t:
                continue
            if any(h in t for h in hints):
                chosen = i
                break
        if chosen is not None:
            used.add(chosen)
        picks.append(chosen)
    return picks


def _guess_kind(titles: list[str]) -> str:
    norm = " ".join(norm_header(t) for t in titles)
    has_meaning = any(h in norm for h in _MEANING_HINTS)
    has_abbrev = any(h in norm for h in _GLOSSARY_KEY_HINTS)
    # A glossary has an abbreviation-style key column AND a meaning column.
    # Anything else (code + definition/description) is treated as a codes file.
    if has_abbrev and has_meaning:
        return "headers"
    return "codes"


def _category_from(path: str, sheet: str, override: Optional[str]) -> str:
    if override:
        return override.strip().lower()
    name = (sheet or "").strip().lower()
    if name and name not in ("sheet1", "sheet", "data"):
        return name
    base = os.path.splitext(os.path.basename(path))[0]
    return base.strip().lower()


def _ingest_headers_sheet(rows: list[list[Any]], lib: Library,
                          report: IngestReport) -> None:
    if not rows:
        return
    hr = _find_header_row(rows)
    titles = [str(c) if c is not None else "" for c in rows[hr]]
    key_i, mean_i, cat_i = _pick_columns(
        titles, _HEADER_KEY_HINTS, _MEANING_HINTS, _CATEGORY_HINTS)
    if key_i is None or mean_i is None:
        # Fall back to first two columns.
        key_i = 0 if key_i is None else key_i
        mean_i = 1 if mean_i is None else mean_i
    for row in rows[hr + 1:]:
        if key_i >= len(row):
            continue
        abbrev = norm_header(row[key_i])
        meaning = str(row[mean_i]).strip() if mean_i < len(row) and row[mean_i] else ""
        if not abbrev or not meaning:
            continue
        category = ""
        if cat_i is not None and cat_i < len(row) and row[cat_i]:
            category = str(row[cat_i]).strip().lower()
        lib.headers[abbrev] = HeaderEntry(
            abbrev=abbrev, meaning=meaning, category=category)
        report.headers_added += 1


def _ingest_codes_sheet(rows: list[list[Any]], lib: Library, category: str,
                        report: IngestReport) -> None:
    if not rows:
        return
    hr = _find_header_row(rows)
    titles = [str(c) if c is not None else "" for c in rows[hr]]
    code_i, def_i = _pick_columns(titles, _CODE_HINTS, _DEF_HINTS)
    if code_i is None:
        code_i = 0
    if def_i is None:
        def_i = 1 if len(titles) > 1 else 0
    # These reference files use a banner header row like
    # "Header : FLD1 | Description of codes of this header". The word "codes"
    # in the description column matches _CODE_HINTS ("CODE") and can steal the
    # code-column pick, collapsing code and definition onto the same column
    # (entries become name->name). When the picks collide, fall back to the
    # conventional first=code / second=definition layout.
    if code_i == def_i and len(titles) > 1:
        code_i, def_i = 0, 1
    cm = lib.code_maps.get(category) or CodeMap(name=category, label=category)
    for row in rows[hr + 1:]:
        if code_i >= len(row):
            continue
        code = norm_code(row[code_i])
        definition = str(row[def_i]).strip() if def_i < len(row) and row[def_i] else ""
        if not code or not definition:
            continue
        cm.entries[code] = definition
        report.codes_added += 1
    cm.reindex()
    lib.code_maps[category] = cm


def ingest_excel(path: str, kind: Optional[str] = None,
                 category: Optional[str] = None,
                 lib: Optional[Library] = None) -> IngestReport:
    """Merge one Excel reference file into the library and persist it.

    ``kind``     -- "headers" or "codes"; auto-detected per sheet when omitted.
    ``category`` -- code domain for a codes file; defaults to the sheet name.
    ``lib``      -- merge into this Library instead of loading from disk (tests).
    """
    own = lib is None
    lib = lib if lib is not None else load_library()
    wb = load_workbook(path, read_only=True, data_only=True)
    report = IngestReport(source=os.path.basename(path), kind=kind or "")
    try:
        for ws in wb.worksheets:
            rows = _rows_of(ws)
            if not any(any(c is not None for c in r) for r in rows):
                continue
            report.sheets.append(ws.title)
            titles = [str(c) if c is not None else ""
                      for c in rows[_find_header_row(rows)]]
            sheet_kind = kind or _guess_kind(titles)
            if sheet_kind == "headers":
                _ingest_headers_sheet(rows, lib, report)
            else:
                cat = _category_from(path, ws.title, category)
                report.category = cat
                _ingest_codes_sheet(rows, lib, cat, report)
    finally:
        wb.close()

    lib.meta.setdefault("sources", [])
    lib.meta["sources"].append(
        {"file": report.source, "kind": report.kind or "auto",
         "category": report.category, "headers": report.headers_added,
         "codes": report.codes_added})
    if own:
        save_library(lib)
    return report
